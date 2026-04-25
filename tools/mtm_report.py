"""
Lector y analizador de archivos MTM para el reporte diario.

Lee los dos Excel del dia desde Google Drive:
- DD-MM-YYYY-CONTROL DIARIO.xlsx
- DD-MM-YYYY-MTM PILAR.xlsm

Y devuelve un analisis completo para mandar por Telegram.
"""

import io
import logging
from datetime import date, datetime

import openpyxl
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

from tools.drive import _get_credentials

log = logging.getLogger(__name__)

# Columnas del CONTROL DIARIO (fila de datos, base 1)
COL_FECHA        = 1   # A
COL_TOTAL_CAJA   = 2   # B
COL_CLIENTES     = 4   # D
COL_STOCK        = 5   # E
COL_TA_CTE_PROV  = 6   # F
COL_RESULTADO    = 7   # G
COL_RESUMEN_DIA  = 8   # H
COL_DIFERENCIA   = 9   # I
COL_GAN_GASTOS   = 10  # J
COL_GAN_MTM      = 11  # K
COL_GASTOS       = 12  # L
COL_TRANSPORTES  = 13  # M
COL_PRODUCCION   = 14  # N
COL_ACT_STOCK    = 15  # O


def _find_file(drive, name_contains: str, fecha_str: str) -> dict | None:
    """Busca un archivo en Drive por nombre exacto del dia."""
    query = f"name contains '{fecha_str}' and name contains '{name_contains}' and trashed = false"
    result = drive.files().list(
        q=query,
        fields="files(id, name)",
        pageSize=5,
        orderBy="modifiedTime desc"
    ).execute()
    files = result.get("files", [])
    return files[0] if files else None


def _download_excel(drive, file_id: str) -> openpyxl.Workbook:
    """Descarga un archivo Excel de Drive y lo abre con openpyxl."""
    request = drive.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True, keep_vba=False)


def _fmt(value) -> str:
    """Formatea un numero como moneda argentina."""
    if value is None:
        return "N/D"
    try:
        n = float(value)
        if abs(n) >= 1_000_000:
            return f"${n/1_000_000:.1f}M"
        elif abs(n) >= 1_000:
            return f"${n/1_000:.0f}K"
        else:
            return f"${n:.2f}"
    except (TypeError, ValueError):
        return str(value)


def _pct(value) -> str:
    if value is None:
        return "N/D"
    try:
        return f"{float(value)*100:.1f}%"
    except (TypeError, ValueError):
        return str(value)


def _get_control_diario_row(ws, fecha_hoy: date):
    """Busca la fila del dia de hoy en el CONTROL DIARIO."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell_fecha = row[COL_FECHA - 1]
        if cell_fecha is None:
            continue
        if isinstance(cell_fecha, datetime):
            if cell_fecha.date() == fecha_hoy:
                return row
        elif isinstance(cell_fecha, date):
            if cell_fecha == fecha_hoy:
                return row
        elif isinstance(cell_fecha, str):
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    if datetime.strptime(cell_fecha.strip(), fmt).date() == fecha_hoy:
                        return row
                except ValueError:
                    pass
    return None


def _analizar_mtm_pilar(wb: openpyxl.Workbook, fecha_hoy: date) -> dict:
    """Extrae datos del archivo MTM PILAR."""
    resultado = {
        "ventas_dia": [],
        "total_ventas": 0,
        "camiones_completos": [],
        "ventas_margen_bajo": [],
        "total_ganancia": 0,
    }

    # Buscar sheet de egresos/ventas (la que tiene columnas de CLIENTE, PRODUCTO, GANANCIA)
    sheet_name = None
    for name in wb.sheetnames:
        if any(kw in name.upper() for kw in ["EGRE", "VENTA", "MTM", "PILAR"]):
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    fecha_str_hoy = fecha_hoy.strftime("%d/%m/%Y")
    fecha_str_hoy2 = fecha_hoy.strftime("%-d/%-m/%Y") if hasattr(fecha_hoy, 'strftime') else fecha_str_hoy

    # Leer encabezados para identificar columnas
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        if any(h and isinstance(h, str) and len(h) > 2 for h in row):
            headers = [str(h).upper().strip() if h else "" for h in row]
            break

    # Indices de columnas clave
    def col_idx(keywords):
        for kw in keywords:
            for i, h in enumerate(headers):
                if kw in h:
                    return i
        return None

    idx_fecha     = col_idx(["FECHA"])
    idx_cliente   = col_idx(["CLIENTE", "PROVEEDOR"])
    idx_producto  = col_idx(["PRODUCTO"])
    idx_cantidad  = col_idx(["CANTIDAD", "CANT"])
    idx_ganancia  = col_idx(["GANANCIA", "RESULT"])
    idx_precio    = col_idx(["PRECIO TOTAL", "TOTAL"])
    idx_costo     = col_idx(["COSTO", "P.COMP", "UNIT"])

    ventas_hoy = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if idx_fecha is None or idx_fecha >= len(row):
            continue
        cell_fecha = row[idx_fecha]
        if cell_fecha is None:
            continue

        es_hoy = False
        if isinstance(cell_fecha, (date, datetime)):
            f = cell_fecha.date() if hasattr(cell_fecha, 'date') else cell_fecha
            es_hoy = (f == fecha_hoy)
        elif isinstance(cell_fecha, str):
            es_hoy = (cell_fecha.strip().startswith(fecha_hoy.strftime("%d/%m")) or
                      cell_fecha.strip().startswith(str(fecha_hoy.day) + "/"))

        if not es_hoy:
            continue

        cantidad = float(row[idx_cantidad]) if idx_cantidad and idx_cantidad < len(row) and row[idx_cantidad] else 0
        ganancia = float(row[idx_ganancia]) if idx_ganancia and idx_ganancia < len(row) and row[idx_ganancia] else 0
        precio   = float(row[idx_precio])   if idx_precio   and idx_precio   < len(row) and row[idx_precio]   else 0
        cliente  = str(row[idx_cliente])    if idx_cliente  and idx_cliente  < len(row) and row[idx_cliente]  else ""
        producto = str(row[idx_producto])   if idx_producto and idx_producto < len(row) and row[idx_producto] else ""

        ventas_hoy.append({
            "cliente":  cliente,
            "producto": producto,
            "cantidad": cantidad,
            "ganancia": ganancia,
            "precio":   precio,
        })

        resultado["total_ganancia"] += ganancia

        # Camion completo: mas de 25000 unidades
        if cantidad >= 25000:
            resultado["camiones_completos"].append({
                "cliente":  cliente,
                "producto": producto,
                "cantidad": int(cantidad),
                "ganancia": ganancia,
            })

        # Margen bajo: menos del 5% (excluyendo negativos intencionales)
        if precio > 0 and ganancia != 0:
            margen = ganancia / precio
            if 0 < margen < 0.05:
                resultado["ventas_margen_bajo"].append({
                    "cliente":  cliente,
                    "producto": producto,
                    "margen":   margen,
                    "ganancia": ganancia,
                })

    resultado["ventas_dia"]   = ventas_hoy
    resultado["total_ventas"] = len(ventas_hoy)
    return resultado


def _analizar_gastos(wb_control: openpyxl.Workbook, fecha_hoy: date) -> list:
    """Busca gastos inusuales (> $30M, no banco/transporte/alquiler)."""
    gastos_alertas = []

    UMBRAL = 30_000_000
    EXCLUIR = ["BANCO", "TRANSPORT", "ALQUIL", "IMPUEST", "RETENCION"]

    # Buscar sheet con gastos detallados
    for sheet_name in wb_control.sheetnames:
        ws = wb_control[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue

            fecha_cell = row[0]
            es_hoy = False
            if isinstance(fecha_cell, (date, datetime)):
                f = fecha_cell.date() if hasattr(fecha_cell, 'date') else fecha_cell
                es_hoy = (f == fecha_hoy)
            elif isinstance(fecha_cell, str):
                es_hoy = fecha_cell.strip().startswith(fecha_hoy.strftime("%d/%m"))

            if not es_hoy:
                continue

            for cell in row[1:]:
                if cell is None:
                    continue
                try:
                    monto = float(cell)
                except (TypeError, ValueError):
                    continue

                if monto > UMBRAL:
                    # Buscar descripcion en la fila
                    desc = " ".join(str(c) for c in row if c and isinstance(c, str))
                    excluir = any(kw in desc.upper() for kw in EXCLUIR)
                    if not excluir:
                        gastos_alertas.append({
                            "descripcion": desc[:80],
                            "monto": monto,
                        })
                    break  # un alerta por fila

    return gastos_alertas


def generar_reporte(fecha: date = None) -> str:
    """
    Genera el reporte diario completo leyendo los archivos MTM del dia.
    Retorna el texto del reporte para mandar por Telegram.
    """
    if fecha is None:
        fecha = date.today()

    fecha_str = fecha.strftime("%d-%m-%Y")
    fecha_display = fecha.strftime("%d/%m/%Y")

    log.info(f"Generando reporte para {fecha_str}...")

    drive = build("drive", "v3", credentials=_get_credentials())

    control_file = _find_file(drive, "CONTROL DIARIO", fecha_str)

    lineas = [f"📊 *Reporte MTM — {fecha_display}*\n"]

    if not control_file:
        return f"📊 Reporte MTM — {fecha_display}\n\nNo se encontro el archivo CONTROL DIARIO del dia."

    # Variables del Excel — se populan si la lectura tiene exito
    ex: dict = {}

    # ── CONTROL DIARIO ────────────────────────────────────────────────────────
    try:
        wb_control = _download_excel(drive, control_file["id"])

        ws = None
        for sheet_name in wb_control.sheetnames:
            s = wb_control[sheet_name]
            header = s.cell(1, 1).value
            if header and "FECHA" in str(header).upper():
                ws = s
                break
        if ws is None:
            ws = wb_control.active

        fila = _get_control_diario_row(ws, fecha)

        if fila:
            ex["diferencia"]  = fila[COL_DIFERENCIA  - 1]
            ex["resultado"]   = fila[COL_RESULTADO    - 1]
            ex["caja"]        = fila[COL_TOTAL_CAJA   - 1]
            ex["clientes"]    = fila[COL_CLIENTES     - 1]
            ex["stock"]       = fila[COL_STOCK        - 1]
            ex["proveedores"] = fila[COL_TA_CTE_PROV  - 1]
            ex["gan_mtm"]     = fila[COL_GAN_MTM      - 1]
            ex["gastos"]      = fila[COL_GASTOS       - 1]
            ex["transportes"] = fila[COL_TRANSPORTES  - 1]

            try:
                dif_val = float(ex["diferencia"]) if ex["diferencia"] else 0
                if abs(dif_val) < 1:
                    lineas.append("✅ *Control cruzado:* CIERRA PERFECTO")
                else:
                    lineas.append(f"⚠️ *Control cruzado:* DIFERENCIA de {_fmt(dif_val)}")
            except (TypeError, ValueError):
                lineas.append("❓ *Control cruzado:* No se pudo leer")

            lineas.append(f"\n💰 *Resultado del dia:* {_fmt(ex['resultado'])}")
            lineas.append(f"🏦 *Total Caja:* {_fmt(ex['caja'])}")
            lineas.append(f"👥 *Clientes (CxC):* {_fmt(ex['clientes'])}")
            lineas.append(f"📦 *Stock valorizado:* {_fmt(ex['stock'])}")
            lineas.append(f"🏭 *Proveedores (CxP):* {_fmt(ex['proveedores'])}")
            lineas.append(f"\n📈 *Ganancia MTM:* {_fmt(ex['gan_mtm'])}")
            lineas.append(f"📉 *Gastos:* {_fmt(ex['gastos'])}")
            lineas.append(f"🚛 *Transportes:* {_fmt(ex['transportes'])}")
        else:
            lineas.append("⚠️ No se encontro la fila del dia en CONTROL DIARIO")

    except Exception as e:
        log.error(f"Error leyendo CONTROL DIARIO: {e}")
        lineas.append(f"⚠️ Error leyendo CONTROL DIARIO: {e}")

    # ── SISTEMA DE GESTIÓN (SQLite) ───────────────────────────────────────────
    datos_sistema = None
    try:
        from tools.sqlite_report import analizar_sistema
        datos_sistema = analizar_sistema(fecha)

        lineas.append(f"\n🖥️ *Sistema de Gestión*")
        lineas.append(f"📦 Ventas: {datos_sistema['ventas_count']} | Total: {_fmt(datos_sistema['ventas_total'])}")
        lineas.append(f"💸 Gastos: {datos_sistema['gastos_count']} | Total: {_fmt(datos_sistema['gastos_total'])}")

        cajas_top = sorted(datos_sistema["cajas"], key=lambda x: abs(x["saldo"]), reverse=True)[:3]
        if cajas_top:
            lineas.append(f"\n🏦 *Cajas principales:*")
            for c in cajas_top:
                lineas.append(f"  • {c['nombre']}: {_fmt(c['saldo'])}")
            lineas.append(f"  *Total: {_fmt(datos_sistema['cajas_total'])}*")

        if datos_sistema["camiones_completos"]:
            lineas.append(f"\n🚛 *Camiones completos ({len(datos_sistema['camiones_completos'])}):*")
            for c in datos_sistema["camiones_completos"][:5]:
                lineas.append(f"  • {c['producto'][:35]} — {c['cantidad']:,} u → {_fmt(c['total'])}")
        else:
            lineas.append("\n🚛 *Camiones completos:* Ninguno")

        if datos_sistema["pedidos"]:
            lineas.append(f"\n📋 *Pedidos del día:*")
            for estado, count in datos_sistema["pedidos"].items():
                lineas.append(f"  • {estado}: {count}")

        if datos_sistema["gastos_alertas"]:
            lineas.append(f"\n⚠️ *Gastos inusuales (>$30M):*")
            for g in datos_sistema["gastos_alertas"][:5]:
                lineas.append(f"  • {g['descripcion'][:50]} — {_fmt(g['monto'])} ({g['categoria']})")

    except Exception as e:
        log.error(f"Error leyendo SQLite: {e}")
        lineas.append(f"\n⚠️ Error leyendo datos del sistema: {e}")

    # ── RECONCILIACIÓN Excel vs SQLite ────────────────────────────────────────
    if ex and datos_sistema:
        try:
            lineas.append(f"\n🔍 *Reconciliación Excel vs Sistema*")

            def _rec(label, excel_val, sqlite_val, umbral_pct=0.02):
                try:
                    e = float(excel_val or 0)
                    s = float(sqlite_val or 0)
                    if e == 0 and s == 0:
                        return f"  ➖ {label}: sin datos en ambos"
                    base = max(abs(e), abs(s), 1)
                    diff = abs(e - s) / base
                    icon = "✅" if diff < umbral_pct else ("⚠️" if diff < 0.10 else "❌")
                    return f"  {icon} {label}: Excel {_fmt(e)} | Sistema {_fmt(s)} | Dif {_fmt(e - s)}"
                except Exception:
                    return f"  ❓ {label}: no se pudo comparar"

            lineas.append(_rec("Gastos",      ex.get("gastos"),      datos_sistema["gastos_total"]))
            lineas.append(_rec("Transportes", ex.get("transportes"), datos_sistema["transportes_total"]))
            lineas.append(_rec("Total caja",  ex.get("caja"),        datos_sistema["cajas_total"]))

        except Exception as e:
            log.error(f"Error en reconciliacion: {e}")

    # ── ANÁLISIS FINANCIERO ───────────────────────────────────────────────────
    if ex:
        try:
            lineas.append(f"\n📐 *Análisis Financiero*")

            caja_v        = float(ex.get("caja")        or 0)
            clientes_v    = float(ex.get("clientes")    or 0)
            stock_v       = float(ex.get("stock")       or 0)
            proveedores_v = float(ex.get("proveedores") or 0)
            resultado_v   = float(ex.get("resultado")   or 0)
            gan_mtm_v     = float(ex.get("gan_mtm")     or 0)
            gastos_v      = float(ex.get("gastos")      or 0)
            transportes_v = float(ex.get("transportes") or 0)

            # Cobertura de deuda: (Caja + CxC) / CxP
            if proveedores_v > 0:
                ratio_cob = (caja_v + clientes_v) / proveedores_v
                icon = "✅" if ratio_cob >= 1.0 else ("⚠️" if ratio_cob >= 0.7 else "❌")
                lineas.append(f"  {icon} Cobertura deuda: {ratio_cob:.2f}x  (Caja+CxC / CxP)")

            # Estructura de activos
            activo_total = caja_v + clientes_v + stock_v
            if activo_total > 0:
                lineas.append(
                    f"  📊 Activos: Caja {caja_v/activo_total*100:.0f}% | "
                    f"CxC {clientes_v/activo_total*100:.0f}% | "
                    f"Stock {stock_v/activo_total*100:.0f}%"
                )

            # Gastos como % de la ganancia
            total_egresos = gastos_v + transportes_v
            if gan_mtm_v > 0:
                ratio_gastos = total_egresos / gan_mtm_v * 100
                icon = "✅" if ratio_gastos < 30 else ("⚠️" if ratio_gastos < 60 else "❌")
                lineas.append(f"  {icon} Gastos+Fletes / Ganancia: {ratio_gastos:.0f}%")

            # Resultado del dia
            if resultado_v >= 0:
                lineas.append(f"  ✅ Resultado del dia: {_fmt(resultado_v)}")
            else:
                lineas.append(f"  ❌ Resultado NEGATIVO: {_fmt(resultado_v)}")

            # Alertas
            if caja_v < proveedores_v * 0.3:
                lineas.append(f"  🚨 *ALERTA LIQUIDEZ:* Caja {_fmt(caja_v)} muy baja vs deuda {_fmt(proveedores_v)}")
            if clientes_v > proveedores_v * 2:
                lineas.append(f"  ⚠️ Alta exposicion clientes: CxC {_fmt(clientes_v)} vs CxP {_fmt(proveedores_v)}")

        except Exception as e:
            log.error(f"Error en analisis financiero: {e}")

    lineas.append(f"\n_Generado automaticamente a las {datetime.now().strftime('%H:%M')}_")

    return "\n".join(lineas)


def generar_reporte_excel(fecha: date = None) -> str:
    """
    Version rapida del reporte — solo lee el CONTROL DIARIO de Drive.
    Sin SQLite. Retorna en ~10-20 segundos.
    """
    if fecha is None:
        fecha = date.today()

    fecha_str = fecha.strftime("%d-%m-%Y")
    fecha_display = fecha.strftime("%d/%m/%Y")

    drive = build("drive", "v3", credentials=_get_credentials())
    control_file = _find_file(drive, "CONTROL DIARIO", fecha_str)

    lineas = [f"📊 *Reporte MTM — {fecha_display}*\n"]

    if not control_file:
        return f"📊 Reporte MTM — {fecha_display}\n\nNo se encontro el archivo CONTROL DIARIO del dia."

    ex: dict = {}

    try:
        wb_control = _download_excel(drive, control_file["id"])
        ws = None
        for sheet_name in wb_control.sheetnames:
            s = wb_control[sheet_name]
            header = s.cell(1, 1).value
            if header and "FECHA" in str(header).upper():
                ws = s
                break
        if ws is None:
            ws = wb_control.active

        fila = _get_control_diario_row(ws, fecha)

        if fila:
            ex["diferencia"]  = fila[COL_DIFERENCIA  - 1]
            ex["resultado"]   = fila[COL_RESULTADO    - 1]
            ex["caja"]        = fila[COL_TOTAL_CAJA   - 1]
            ex["clientes"]    = fila[COL_CLIENTES     - 1]
            ex["stock"]       = fila[COL_STOCK        - 1]
            ex["proveedores"] = fila[COL_TA_CTE_PROV  - 1]
            ex["gan_mtm"]     = fila[COL_GAN_MTM      - 1]
            ex["gastos"]      = fila[COL_GASTOS       - 1]
            ex["transportes"] = fila[COL_TRANSPORTES  - 1]

            try:
                dif_val = float(ex["diferencia"]) if ex["diferencia"] else 0
                lineas.append("✅ *Control cruzado:* CIERRA PERFECTO" if abs(dif_val) < 1
                              else f"⚠️ *Control cruzado:* DIFERENCIA de {_fmt(dif_val)}")
            except (TypeError, ValueError):
                lineas.append("❓ *Control cruzado:* No se pudo leer")

            lineas.append(f"\n💰 *Resultado del dia:* {_fmt(ex['resultado'])}")
            lineas.append(f"🏦 *Total Caja:* {_fmt(ex['caja'])}")
            lineas.append(f"👥 *Clientes (CxC):* {_fmt(ex['clientes'])}")
            lineas.append(f"📦 *Stock valorizado:* {_fmt(ex['stock'])}")
            lineas.append(f"🏭 *Proveedores (CxP):* {_fmt(ex['proveedores'])}")
            lineas.append(f"\n📈 *Ganancia MTM:* {_fmt(ex['gan_mtm'])}")
            lineas.append(f"📉 *Gastos:* {_fmt(ex['gastos'])}")
            lineas.append(f"🚛 *Transportes:* {_fmt(ex['transportes'])}")
        else:
            lineas.append("⚠️ No se encontro la fila del dia en CONTROL DIARIO")

    except Exception as e:
        log.error(f"Error leyendo CONTROL DIARIO: {e}")
        lineas.append(f"⚠️ Error leyendo CONTROL DIARIO: {e}")

    # Analisis financiero solo con datos del Excel
    if ex:
        try:
            caja_v        = float(ex.get("caja")        or 0)
            clientes_v    = float(ex.get("clientes")    or 0)
            stock_v       = float(ex.get("stock")       or 0)
            proveedores_v = float(ex.get("proveedores") or 0)
            resultado_v   = float(ex.get("resultado")   or 0)
            gan_mtm_v     = float(ex.get("gan_mtm")     or 0)
            gastos_v      = float(ex.get("gastos")      or 0)
            transportes_v = float(ex.get("transportes") or 0)

            lineas.append(f"\n📐 *Análisis Financiero*")

            if proveedores_v > 0:
                ratio_cob = (caja_v + clientes_v) / proveedores_v
                icon = "✅" if ratio_cob >= 1.0 else ("⚠️" if ratio_cob >= 0.7 else "❌")
                lineas.append(f"  {icon} Cobertura deuda: {ratio_cob:.2f}x  (Caja+CxC / CxP)")

            activo_total = caja_v + clientes_v + stock_v
            if activo_total > 0:
                lineas.append(
                    f"  📊 Activos: Caja {caja_v/activo_total*100:.0f}% | "
                    f"CxC {clientes_v/activo_total*100:.0f}% | "
                    f"Stock {stock_v/activo_total*100:.0f}%"
                )

            total_egresos = gastos_v + transportes_v
            if gan_mtm_v > 0:
                ratio_gastos = total_egresos / gan_mtm_v * 100
                icon = "✅" if ratio_gastos < 30 else ("⚠️" if ratio_gastos < 60 else "❌")
                lineas.append(f"  {icon} Gastos+Fletes / Ganancia: {ratio_gastos:.0f}%")

            lineas.append(f"  {'✅' if resultado_v >= 0 else '❌'} Resultado: {_fmt(resultado_v)}")

            if caja_v < proveedores_v * 0.3:
                lineas.append(f"  🚨 *ALERTA LIQUIDEZ:* Caja {_fmt(caja_v)} muy baja vs deuda {_fmt(proveedores_v)}")
            if clientes_v > proveedores_v * 2:
                lineas.append(f"  ⚠️ Alta exposicion clientes: CxC {_fmt(clientes_v)} vs CxP {_fmt(proveedores_v)}")

        except Exception as e:
            log.error(f"Error en analisis financiero: {e}")

    lineas.append(f"\n_Generado a las {datetime.now().strftime('%H:%M')} — datos del sistema en camino..._")
    return "\n".join(lineas)
